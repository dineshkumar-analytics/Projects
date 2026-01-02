import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile
import time

st.set_page_config(page_title="Team Comparison Tool", layout="wide")
st.title("Compare TEAM A & TEAM B Between Two Excel Files")

old_file = st.file_uploader("Upload OLD Excel file", type=["xlsx"])
new_file = st.file_uploader("Upload NEW Excel file", type=["xlsx"])

if old_file and new_file:

    old_df = pd.read_excel(old_file, usecols=[0,1], header=None)
    new_df = pd.read_excel(new_file, usecols=[0,1], header=None)

    old_df.columns = ["TEAM A_OLD", "TEAM B_OLD"]
    new_df.columns = ["TEAM A_NEW", "TEAM B_NEW"]

    max_len = max(len(old_df), len(new_df))
    old_df = old_df.reindex(range(max_len))
    new_df = new_df.reindex(range(max_len))

    combined = pd.concat([old_df, new_df], axis=1).fillna("")

    diff_mask = pd.DataFrame(False, index=combined.index, columns=combined.columns)

    for old_col, new_col in [
        ("TEAM A_OLD","TEAM A_NEW"),
        ("TEAM B_OLD","TEAM B_NEW")
    ]:
        diff = combined[old_col].astype(str).str.strip() != combined[new_col].astype(str).str.strip()
        diff_mask[old_col] = diff
        diff_mask[new_col] = diff

    st.subheader("Comparison Result")
    st.dataframe(combined, use_container_width=True)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    output = tmp.name
    combined.to_excel(output, index=False, header=False)

    wb = load_workbook(output)
    ws = wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r in range(1, max_len + 1):
        for c in range(1, len(combined.columns) + 1):
            if diff_mask.iloc[r-1, c-1]:
                ws.cell(row=r, column=c).fill = yellow

    wb.save(output)

    with open(output, "rb") as f:
        st.download_button("Download Highlighted Excel", f,
                           file_name=f"teams_difference_{int(time.time())}.xlsx")

    st.success("✅ Done — differences highlighted in Excel")

else:
    st.info("Please upload both OLD and NEW Excel files.")
